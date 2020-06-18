#!/usr/bin/python3
# @Author: WLP
# @name: openpyxl操作.py
# @date 2020-04-26 16:34

#说明文档https://openpyxl.readthedocs.io/en/stable/#

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

#创建新的文件
wb = Workbook()
dest_filename = 'empty_book.xlsx'
ws1 = wb.active
ws1.title = "range names"
for row in range(1, 40):
    ws1.append(range(600))
ws2 = wb.create_sheet(title="Pi")
ws2['F5'] = 3.14
ws3 = wb.create_sheet(title="Data")
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)

wb.save(filename = dest_filename)

from openpyxl import load_workbook

#打开已有文件
wb = load_workbook(filename = 'empty_book.xlsx')
sheet_ranges = wb['range names']
print(sheet_ranges['A1'].value)#获取已有文件的值
sheet_ranges['A1'] = 123#修改已有文件的值
wb.save('empty_book.xlsx')



